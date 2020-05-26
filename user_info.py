import os
import django
from openpyxl import Workbook, load_workbook
import yagmail


os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'curiousweb.settings')
django.setup()

from main_app.models import User


def find():

    file = 'RY-2020-ZOOM-LIST-edited.xlsx'
    wb = load_workbook(file)
    sheetnames = wb.sheetnames #this is a list of sheetnames
    ws = wb[sheetnames[0]]

    username_col = 6
    password_col = 7
    email_col = 10

    rowcount = 1

    for row in ws.rows:
        if str(row[username_col].value) == 'end':
            wb.save(file)
            return

        print(row[username_col].value)
        try:
            user = User.objects.get(username = str(row[username_col].value))
        except:
            pass

        try:
            ws.cell(row=rowcount, column=10).value = user.email
        except:
            pass


        rowcount = rowcount + 1



if __name__ == '__main__':
    print('running user_info.py...')
    print()
    find()
    print('done.')
