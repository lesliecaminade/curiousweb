import os
import django
from openpyxl import Workbook, load_workbook
import yagmail


os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'curiousweb.settings')
django.setup()

from main_app.models import User


def find_and_send():

    file = 'batch2-cleaned.xlsx'
    wb = load_workbook(file, read_only = True)
    #sheetnames = wb.sheetnames #this is a list of sheetnames
    ws = wb.active

    username_col = 6
    password_col = 7

    for row in ws.rows:
        print(row[username_col].value)
        user = User.objects.get(username = str(row[username_col].value))
        first_name = user.first_name
        username = user.username
        password = str(row[password_col])
        params = {
            'first_name': user.first_name,
            'email': user.email,
            'username': user.username,
            'password': str(row[password_col].value),
        }
        send_email(params)


    file.close()



def send_email(dict):

    """yagmail is a library to manage google smtp in a more simpler manner,
    for more information, visit https://github.com/kootenpv/yagmail"""

    yag = yagmail.SMTP('cortexsilicon','jnzbhrbqcsavnlhu') #input the email username and app password
    contents = f"""Greetings {dict['first_name']},

We have received your registration and your payment and we gladly inform you that you are now officially enrolled in CERTC’s Online Review Platform which will start on Monday next week May 25,2020.

Thank you for joining CERTC’s online review program.

Here are your log-in credentials in our site certconlinereview.com which you will use throughout this course.

User Name: {dict['username']}
Password: {dict['password']}
Note: Full access to our website will start on the second week of our cloud classes (Jun 1,2020).

Once again thank you! If you have any questions, please let us know!

Facebook:
Jiovanni Quiseo
CERTC Review Center
Contact Nos.: 09173028824/09321751218
    """

    yag.send(to = ['jmquiseo@gmail.com', dict['email']], subject = 'CERTC Enrollment', contents = contents) #send the email
    print('sent email to', dict['email'])

    yag.send(to = ['lesliecaminade@gmail.com'], subject = 'CERTC Enrollment', contents = contents) #send the email





if __name__ == '__main__':
    print('running batchemail.py...')
    print()
    find_and_send()
    print('done.')
