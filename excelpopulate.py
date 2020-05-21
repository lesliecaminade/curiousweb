from openpyxl import Workbook, load_workbook
from faker import Faker

fake = Faker()

def main():
    wb = Workbook()
    ws = wb.active


    number_of_fake_mcqs = 100
    for i in range(1, number_of_fake_mcqs):
        print(fake.paragraph())
        ws['A'+str(i)] = fake.paragraph()
        ws['B'+str(i)] = fake.sentence()
        ws['C'+str(i)] = fake.sentence()
        ws['D'+str(i)] = fake.sentence()
        ws['E'+str(i)] = fake.sentence()


    wb.save(filename = 'test.xlsx')


if __name__ == '__main__':
    print('running excelpopulate.py....')
    main()
    print('done.')
