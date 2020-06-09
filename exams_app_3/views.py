from django.shortcuts import render
from django.urls import reverse_lazy, reverse
from django.http import HttpResponse, Http404
from django.views.generic import (View,TemplateView,
                                ListView,DetailView,
                                DeleteView, CreateView,
                                UpdateView)
from . import models
from . import forms

from django.contrib import messages
from django.contrib.auth.mixins import(
    LoginRequiredMixin,
    PermissionRequiredMixin
)
from django.shortcuts import get_object_or_404
from datetime import datetime
from django.http import HttpResponse, HttpResponseRedirect
from django.utils.timezone import make_aware
from openpyxl import Workbook
from openpyxl import load_workbook

class ExamView(View):
    def get(self, *args, **kwargs):
        exam = models.Exam.objects.get(pk = int(self.kwargs.get('pk')))

        if self.request.user.is_authenticated:
            if exam.is_accessible or self.request.user.is_superuser:
                #try to find first question that were never accessed by you
                items = exam.items.exclude(access_count__user = self.request.user).order_by('?')
                try:
                    if not items.exists():
                        #if none, find the item that has the least number of access times
                        print('No item unaccessed,')
                        items = exam.items.filter(access_count__user = self.request.user).order_by('access_count__count', '?')
                        item = items[0]
                        count_old = item.access_count.filter(user = self.request.user)[0].count
                        print(count_old)
                        item.access_count.filter(user = self.request.user).update(count = count_old + 1)
                        print(item.access_count.filter(user = self.request.user)[0].count)
                    else:
                        item = items[0]
                        new_mcq_access_count = models.MCQAccessCount.objects.create(
                            user = self.request.user,
                            count = 1,
                        )
                        new_mcq_access_count.save()
                        item.access_count.add(new_mcq_access_count)
                except:
                    return HttpResponseRedirect(reverse('index', kwargs = {'activetab': 'problemsolving'}))


                template_name = 'exams_app_3/exam_detail.html'
                context = {
                    'exam': exam,
                    'item': item,
                    'access_count': item.access_count.filter(user = self.request.user)[0],
                    }
                return render(self.request, template_name, context)

class CreateExamManual(View):
    def get(self, *args, **kwargs):
        # categoryapk = int(self.kwargs['pk'])
        template_name = 'exams_app_3/create_exam_manual.html'
        context = {
            # 'categoryapk': categoryapk,
        }
        return render(self.request, template_name, context)

    def post(self, *args, **kwargs):
        exam = models.Exam.objects.create(
            title = str(self.request.POST.get('title')),
            description = str(self.request.POST.get('description')),
            is_ece = bool(self.request.POST.get('is_ece')),
            is_ee = bool(self.request.POST.get('is_ee')),
            is_tutorial = bool(self.request.POST.get('is_tutorial')),
            is_accessible = bool(self.request.POST.get('is_accessible')),
        )
        exam.save()

        return HttpResponseRedirect(reverse('exams_app_3:create_exam_manual_add_item', kwargs = {'pk': exam.pk}))

class CreateExamManualAddItem(View):
    def get(self, *args, **kwargs):
        template_name = 'exams_app_3/exam_form.html'
        exampk = int(self.kwargs['pk'])
        exam = models.Exam.objects.get(pk = exampk)

        context = {
            'exampk': exampk,
            'exam': exam,
        }
        return render(self.request, template_name, context)

    def post(self, *args, **kwargs):
        exampk = int(self.request.POST.get('pk'))
        question = models.MCQ.objects.create(
            question = self.request.POST.get('question'),
            image = self.request.FILES.get('image'),
            explanation_image = self.request.FILES.get('explanation_image'),
            )
        question.save()

        letters = ['a', 'b', 'c', 'd']
        for letter in letters:
            choice = models.Choice.objects.create(
                content = self.request.POST.get('choice_' + letter),
                correct = bool(self.request.POST.get('choice_' + letter + '_correct')),
                image = self.request.FILES.get('choice_' + letter + '_image'),
            )
            choice.save()
            question.choices.add(choice)

        exam = models.Exam.objects.get(pk = exampk)
        exam.items.add(question)

        return HttpResponseRedirect(reverse('exams_app_3:create_exam_manual_add_item', kwargs = {'pk': exampk,}))


class DeleteItemView(View):
    def get(self, *args, **kwargs):
        itempk = int(self.kwargs['itempk'])
        categoryapk = int(self.kwargs['categoryapk'])
        exampk = int(self.kwargs['exampk'])

        item = models.MCQ.objects.get(pk = itempk)
        item.delete()

        return HttpResponseRedirect(reverse('exams_app_3:update_exam', kwargs = {
            'categoryapk': categoryapk,
            'exampk': exampk,
        }))

class DeleteExamView(DeleteView):
    template_name = 'exams_app/confirm_delete.html'
    model = models.Exam
    success_url = reverse_lazy('index', kwargs = {'activetab': 'problemsolving'})
    extra_context = {'nav_exams': 'active',}

class CreateExamUploadView(View):
    def get(self, *args, **kwargs):
        template_name = 'exams_app_3/exam_create_upload.html'
        return render(self.request, template_name)

    def post(self, *args, **kwarg):
        file = self.request.FILES.get('examfile')

        if not file:
            return render(self.request, 'exams_app_3/alerts.html', {'message': 'Please upload a file', 'type': 'warning'})

        if not self.request.POST.get('title'):
            return render(self.request, 'exams_app_3/alerts.html', {'message': 'Please put a title', 'type': 'warning'})

        new_exam = models.Exam(
            title = self.request.POST.get('title'),
            description = self.request.POST.get('description'),
            is_ece = bool(self.request.POST.get('is_ece')),
            is_ee = bool(self.request.POST.get('is_ee')),
            is_tutorial = bool(self.request.POST.get('is_tutorial')),
            is_accessible = bool(self.request.POST.get('is_accessible'))
        )
        new_exam.save()

        wb = load_workbook(file, read_only = True)
        sheetnames = wb.sheetnames #this is a list of sheetnames
        ws = wb[sheetnames[0]] #ill just take the first whatever the first sheet is

        mcqs = [] #this will be the container for the mcqs for the bulk create later

        for row in ws.rows:
            new_mcq = models.MCQ(
                question = row[0].value,
                explanation = row[5].value,
            )
            new_mcq.save()

            for col in range(1, 5):
                if '##' in row[col].value:
                    correct = True
                else:
                    correct = False

                new_choice = models.Choice(
                    content = row[col].value.replace('##', ''),
                    correct = correct,
                )
                new_choice.save()
                new_mcq.choices.add(new_choice)
            new_exam.items.add(new_mcq)

        return HttpResponseRedirect(reverse('index', kwargs = {'activetab': 'problemsolving',}))

class MCQList(View):
    def get(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_superuser:
                exams = models.Exam.objects.all().order_by('-pk')
            elif self.request.user.is_ece:
                exams = models.Exam.objects.filter(is_ece = True).order_by('-pk')
            elif self.request.user.is_ee:
                exams = models.Exam.objects.filter(is_ee = True).order_by('-pk')
            elif self.request.user.is_tutorial:
                exams = models.Exam.objects.filter(is_tutorial = True).order_by('-pk')

            template_name = 'exams_app_3/exam_list.html'
            context = {
                'exams': exams,
            }

            return render(self.request, template_name, context)

class LockExamView(View):
    def get(self, *args, **kwargs):
        if self.request.user.is_superuser:
            exam = models.Exam.objects.filter(pk = int(self.kwargs.get('pk'))).update(is_accessible = False)
            return HttpResponseRedirect(reverse('index', kwargs = {'activetab': 'problemsolving',}))

class UnlockExamView(View):
    def get(self, *args, **kwargs):
        if self.request.user.is_superuser:
            exam = models.Exam.objects.filter(pk = int(self.kwargs.get('pk'))).update(is_accessible = True)
            return HttpResponseRedirect(reverse('index', kwargs = {'activetab': 'problemsolving',}))

class ToggleFlag(View):
    def get(self, *args, **kwargs):
        if self.request.user.is_superuser:
            exam = models.Exam.objects.filter(pk = int(self.kwargs.get('pk')))
            flag = self.kwargs.get('flag')
            if self.kwargs.get('flag') == 'is_ece':
                exam.update(is_ece = bool(self.kwargs.get('setting', False)))
            elif self.kwargs.get('flag') == 'is_ee':
                exam.update(is_ee = bool(self.kwargs.get('setting', False)))
            elif self.kwargs.get('flag') == 'is_tutorial':
                exam.update(is_tutorial = bool(self.kwargs.get('setting', False)))

            return HttpResponseRedirect(reverse('index', kwargs = {'activetab': 'problemsolving'}))
